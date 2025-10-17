"""
Report Generation Service - Phase 3 Day 7
==========================================
Generates PDF and Excel reports for activity and performance data.
"""
import logging
from datetime import datetime, date, timedelta
from typing import Dict, Any, List, Optional
from io import BytesIO
import json

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Service for generating reports in various formats"""
    
    def __init__(self):
        pass
    
    def generate_activity_report(
        self,
        data: Dict[str, Any],
        format: str = 'json'
    ) -> bytes:
        """
        Generate activity report in specified format.
        
        Args:
            data: Report data
            format: Output format ('json', 'csv', 'pdf', 'excel')
        
        Returns:
            Report content as bytes
        """
        if format == 'json':
            return self._generate_json_report(data)
        elif format == 'csv':
            return self._generate_csv_report(data)
        elif format == 'pdf':
            return self._generate_pdf_report(data)
        elif format == 'excel':
            return self._generate_excel_report(data)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_json_report(self, data: Dict[str, Any]) -> bytes:
        """Generate JSON report"""
        return json.dumps(data, indent=2, default=str).encode('utf-8')
    
    def _generate_csv_report(self, data: Dict[str, Any]) -> bytes:
        """Generate CSV report"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Metric', 'Value'])
        
        # Write data
        for key, value in data.items():
            if isinstance(value, (str, int, float)):
                writer.writerow([key, value])
        
        return output.getvalue().encode('utf-8')
    
    def _generate_pdf_report(self, data: Dict[str, Any]) -> bytes:
        """
        Generate PDF report.
        Note: Requires reportlab library
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4f46e5'),
                spaceAfter=30,
            )
            story.append(Paragraph("Activity Report", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Report date
            date_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            story.append(Paragraph(date_text, styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Summary section
            story.append(Paragraph("Summary", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            # Create table data
            table_data = [['Metric', 'Value']]
            for key, value in data.items():
                if isinstance(value, (str, int, float)):
                    table_data.append([key.replace('_', ' ').title(), str(value)])
            
            # Create table
            table = Table(table_data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            
            return buffer.getvalue()
            
        except ImportError:
            logger.warning("reportlab not installed, returning JSON instead")
            return self._generate_json_report(data)
    
    def _generate_excel_report(self, data: Dict[str, Any]) -> bytes:
        """
        Generate Excel report.
        Note: Requires openpyxl library
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Activity Report"
            
            # Header styling
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Write header
            ws['A1'] = 'Metric'
            ws['B1'] = 'Value'
            ws['A1'].fill = header_fill
            ws['B1'].fill = header_fill
            ws['A1'].font = header_font
            ws['B1'].font = header_font
            
            # Write data
            row = 2
            for key, value in data.items():
                if isinstance(value, (str, int, float)):
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = value
                    row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            
            return buffer.getvalue()
            
        except ImportError:
            logger.warning("openpyxl not installed, returning CSV instead")
            return self._generate_csv_report(data)
