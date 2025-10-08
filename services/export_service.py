"""
Export Service - Export candidate data to CSV and Excel formats
"""
import csv
import io
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting candidate data to various formats"""
    
    def export_to_csv(self, candidates: List[Dict[str, Any]], include_scores: bool = True) -> str:
        """
        Export candidates to CSV format.
        
        Args:
            candidates: List of candidate dictionaries
            include_scores: Whether to include authenticity/match scores
            
        Returns:
            CSV string
        """
        if not candidates:
            return ""
        
        # Create CSV in memory
        output = io.StringIO()
        
        # Define columns
        columns = [
            'ID', 'Name', 'Email', 'Phone', 'LinkedIn', 
            'Skills', 'Experience (Years)', 'Education', 
            'Status', 'Resume Count', 'Created At'
        ]
        
        if include_scores:
            columns.extend(['Authenticity Score', 'Match Score'])
        
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        
        # Write data
        for candidate in candidates:
            row = {
                'ID': candidate.get('id', ''),
                'Name': candidate.get('name', ''),
                'Email': candidate.get('email', ''),
                'Phone': candidate.get('phone', ''),
                'LinkedIn': candidate.get('linkedin', ''),
                'Skills': ', '.join(candidate.get('skills', [])),
                'Experience (Years)': candidate.get('experience_years', 0),
                'Education': candidate.get('education', ''),
                'Status': candidate.get('status', ''),
                'Resume Count': candidate.get('resume_count', 0),
                'Created At': candidate.get('created_at', '')
            }
            
            if include_scores:
                row['Authenticity Score'] = candidate.get('authenticity_score', '')
                row['Match Score'] = candidate.get('match_score', '')
            
            writer.writerow(row)
        
        return output.getvalue()
    
    def export_to_excel(self, candidates: List[Dict[str, Any]], 
                       include_scores: bool = True,
                       sheet_name: str = "Candidates") -> bytes:
        """
        Export candidates to Excel format with formatting.
        
        Args:
            candidates: List of candidate dictionaries
            include_scores: Whether to include authenticity/match scores
            sheet_name: Name of the Excel sheet
            
        Returns:
            Excel file as bytes
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Define columns
        columns = [
            'ID', 'Name', 'Email', 'Phone', 'LinkedIn', 
            'Skills', 'Experience (Years)', 'Education', 
            'Status', 'Resume Count', 'Created At'
        ]
        
        if include_scores:
            columns.extend(['Authenticity Score', 'Match Score'])
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_num, candidate in enumerate(candidates, 2):
            data = [
                candidate.get('id', ''),
                candidate.get('name', ''),
                candidate.get('email', ''),
                candidate.get('phone', ''),
                candidate.get('linkedin', ''),
                ', '.join(candidate.get('skills', [])),
                candidate.get('experience_years', 0),
                candidate.get('education', ''),
                candidate.get('status', ''),
                candidate.get('resume_count', 0),
                candidate.get('created_at', '')
            ]
            
            if include_scores:
                data.extend([
                    candidate.get('authenticity_score', ''),
                    candidate.get('match_score', '')
                ])
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                
                # Center align numeric columns
                if col_num in [1, 7, 10]:  # ID, Experience, Resume Count
                    cell.alignment = Alignment(horizontal="center")
                
                # Color code status
                if col_num == 9:  # Status column
                    if value == "Hired":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "Rejected":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == "Interviewed":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Auto-adjust column widths
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Set minimum width based on column type
            if column_title == 'Skills':
                ws.column_dimensions[column_letter].width = 40
            elif column_title in ['Email', 'LinkedIn']:
                ws.column_dimensions[column_letter].width = 30
            elif column_title == 'Name':
                ws.column_dimensions[column_letter].width = 25
            elif column_title in ['Phone', 'Created At']:
                ws.column_dimensions[column_letter].width = 20
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add metadata sheet
        meta_ws = wb.create_sheet("Export Info")
        meta_ws['A1'] = "Export Date"
        meta_ws['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_ws['A2'] = "Total Candidates"
        meta_ws['B2'] = len(candidates)
        meta_ws['A3'] = "Exported By"
        meta_ws['B3'] = "AI HR Assistant"
        
        # Style metadata
        for row in meta_ws['A1:B3']:
            for cell in row:
                cell.font = Font(bold=True if cell.column == 1 else False)
                cell.border = border
        
        meta_ws.column_dimensions['A'].width = 20
        meta_ws.column_dimensions['B'].width = 30
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_detailed_excel(self, candidates: List[Dict[str, Any]], 
                            detailed_data: Dict[int, Dict[str, Any]] = None) -> bytes:
        """
        Export candidates with detailed information in separate sheets.
        
        Args:
            candidates: List of candidate dictionaries
            detailed_data: Optional dict mapping candidate_id to detailed info
                          (education, work_experience, certifications, etc.)
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Main candidates sheet
        ws_main = wb.active
        ws_main.title = "Candidates Summary"
        
        # Use the standard export for main sheet
        temp_output = io.BytesIO()
        temp_wb = Workbook()
        temp_ws = temp_wb.active
        
        # Copy data from standard export
        standard_bytes = self.export_to_excel(candidates, include_scores=True)
        temp_wb = Workbook()
        temp_wb = openpyxl.load_workbook(io.BytesIO(standard_bytes))
        
        # If detailed data provided, create additional sheets
        if detailed_data:
            # Education sheet
            ws_edu = wb.create_sheet("Education Details")
            edu_headers = ['Candidate ID', 'Name', 'Degree', 'Field of Study', 'Institution', 'GPA', 'Graduation Year']
            for col_num, header in enumerate(edu_headers, 1):
                ws_edu.cell(row=1, column=col_num, value=header).font = Font(bold=True)
            
            # Work Experience sheet
            ws_exp = wb.create_sheet("Work Experience")
            exp_headers = ['Candidate ID', 'Name', 'Company', 'Title', 'Location', 'Start Date', 'End Date', 'Duration (Years)']
            for col_num, header in enumerate(exp_headers, 1):
                ws_exp.cell(row=1, column=col_num, value=header).font = Font(bold=True)
            
            # Certifications sheet
            ws_cert = wb.create_sheet("Certifications")
            cert_headers = ['Candidate ID', 'Name', 'Certification', 'Issuing Organization', 'Year']
            for col_num, header in enumerate(cert_headers, 1):
                ws_cert.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
